from random import randint
from openpyxl import Workbook, load_workbook

def randNumPopulate():

    wb = Workbook()
    destFilename = 'testBook.xlsx'

    ws = wb.active
    
    for x in range(1,13):
        for y in range(1,13):
            number = randint(1, 1000)
            ws.cell(column = x, row = y, value = number)

    wb.save(filename = destFilename)
    return 1